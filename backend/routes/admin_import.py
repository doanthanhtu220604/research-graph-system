"""
Admin API - Import dữ liệu từ file Excel/CSV
Hỗ trợ import: Giảng viên, Công trình, Đề tài, Bộ môn
"""

import io
import math
import pandas as pd
from flask import Blueprint, jsonify, request
from backend.services.neo4j_connection import get_neo4j_connection, generate_slug

admin_import_bp = Blueprint("admin_import_api", __name__)

# ─────────────────────────────────────────────
#  Helper utilities
# ─────────────────────────────────────────────

def safe_str(val) -> str:
    """Chuyển giá trị thành string sạch, bỏ NaN/None."""
    if val is None:
        return ""
    if isinstance(val, float) and math.isnan(val):
        return ""
    return str(val).strip()


def parse_list_field(val) -> list[str]:
    """Chuyển chuỗi phân cách bởi dấu '|' hoặc ',' thành list."""
    s = safe_str(val)
    if not s:
        return []
    sep = "|" if "|" in s else ","
    return [x.strip() for x in s.split(sep) if x.strip()]


def read_file_to_df(file_storage) -> pd.DataFrame:
    """Đọc file upload (xlsx/csv) thành DataFrame."""
    filename = file_storage.filename.lower()
    content = file_storage.read()
    if filename.endswith(".csv"):
        # Thử utf-8, fallback utf-8-sig (Excel VN)
        try:
            df = pd.read_csv(io.BytesIO(content), encoding="utf-8-sig", dtype=str)
        except Exception:
            df = pd.read_csv(io.BytesIO(content), encoding="cp1252", dtype=str)
    else:
        df = pd.read_excel(io.BytesIO(content), dtype=str)

    # Chuẩn hóa tên cột: strip spaces, lower
    df.columns = [c.strip() for c in df.columns]
    # Bỏ các dòng hoàn toàn rỗng
    df.dropna(how="all", inplace=True)
    return df


# ─────────────────────────────────────────────
#  Import theo loại
# ─────────────────────────────────────────────

def import_giang_vien(df: pd.DataFrame, conn) -> dict:
    """
    Cột bắt buộc: ho_va_ten
    Cột tuỳ chọn: ma_gv, hoc_vi, chuc_danh, chuc_vu, email,
                  dien_thoai, chuyen_nganh, trang_thai_cong_tac,
                  bo_mon, linh_vuc_nghien_cuu (phân cách '|')
    """
    created, updated, errors = 0, 0, []

    rows = df.to_dict(orient="records")
    for idx, row in enumerate(rows, start=2):   # start=2 vì dòng 1 là header
        ho_va_ten = safe_str(row.get("ho_va_ten"))
        if not ho_va_ten:
            errors.append(f"Dòng {idx}: thiếu họ và tên – bỏ qua.")
            continue

        ho_va_ten = ho_va_ten.upper()
        props = {
            "ho_va_ten": ho_va_ten,
            "ma_gv":                safe_str(row.get("ma_gv")),
            "hoc_vi":               safe_str(row.get("hoc_vi")).upper() if safe_str(row.get("hoc_vi")) else None,
            "chuc_danh":            safe_str(row.get("chuc_danh")).upper() if safe_str(row.get("chuc_danh")) else None,
            "chuc_vu":              safe_str(row.get("chuc_vu")).upper() if safe_str(row.get("chuc_vu")) else None,
            "email":                safe_str(row.get("email")),
            "dien_thoai":           safe_str(row.get("dien_thoai")),
            "chuyen_nganh":         safe_str(row.get("chuyen_nganh")).upper() if safe_str(row.get("chuyen_nganh")) else None,
            "trang_thai_cong_tac":  safe_str(row.get("trang_thai_cong_tac")) or "Đang công tác",
            "anh_dai_dien":         safe_str(row.get("anh_dai_dien")),
        }

        try:
            # Tìm giảng viên đã tồn tại bằng email hoặc ma_gv (nếu có), hoặc ho_va_ten làm fallback
            email = props["email"]
            ma_gv = props["ma_gv"]
            ho_va_ten = props["ho_va_ten"]

            existing = None
            if email or ma_gv:
                existing = conn.query_single("""
                    MATCH (gv:GiangVien)
                    WHERE ((gv.email = $email AND $email <> "")
                       OR (gv.ma_gv = $ma_gv AND $ma_gv <> ""))
                      AND coalesce(gv.is_deleted, false) = false
                    RETURN gv.id AS gv_id
                """, {"email": email, "ma_gv": ma_gv})

            if not existing and ho_va_ten:
                existing = conn.query_single("""
                    MATCH (gv:GiangVien)
                    WHERE toUpper(gv.ho_va_ten) = toUpper($ho_va_ten)
                      AND coalesce(gv.is_deleted, false) = false
                    RETURN gv.id AS gv_id
                """, {"ho_va_ten": ho_va_ten})

            if existing:
                errors.append(f"Dòng {idx}: Giảng viên '{ho_va_ten}' đã tồn tại trong hệ thống (trùng lặp).")
                continue

            result = conn.write("""
                CREATE (gv:GiangVien)
                SET gv.id = 'gv_' + toString(id(gv)),
                    gv.created_at = timestamp(),
                    gv.is_deleted = false,
                    gv += $props
                RETURN gv.id AS gv_id
            """, {"props": props})
            gv_id = result[0]["gv_id"]
            created += 1

            # Trạng thái chuyển công tác -> Chuyển thành Tác giả ngoài
            if props.get("trang_thai_cong_tac") == "Chuyển công tác" and gv_id:
                new_id = f"tgn_{gv_id}"
                conn.write("""
                    MATCH (g:GiangVien {id: $id})
                    OPTIONAL MATCH (g)-[r:THUOC_BO_MON]->()
                    DELETE r
                    WITH g
                    REMOVE g:GiangVien
                    SET g:TacGiaNgoai
                    SET g.id = $new_id,
                        g.don_vi_cong_tac = 'GIẢNG VIÊN ĐÃ CHUYỂN ĐI CỦA TRƯỜNG ĐẠI HỌC NHA TRANG',
                        g.trang_thai = 'Đã duyệt',
                        g.is_deleted = false
                    REMOVE g.password, g.trang_thai_cong_tac, g.chuc_vu, g.ma_gv, g.dien_thoai
                """, {"id": gv_id, "new_id": new_id})
                continue

            # Xử lý Bộ môn
            bo_mon = safe_str(row.get("bo_mon"))
            if bo_mon:
                bo_mon = bo_mon.upper()
                conn.write("""
                    MATCH (gv:GiangVien) WHERE gv.id = $gv_id
                    MERGE (bm:BoMon {ten_bo_mon: $bo_mon})
                    ON CREATE SET bm.id = 'bm_' + toString(id(bm)), bm.is_deleted = false
                    MERGE (gv)-[:THUOC_BO_MON]->(bm)
                """, {"gv_id": gv_id, "bo_mon": bo_mon})

            # Xử lý Lĩnh vực nghiên cứu
            linh_vucs = parse_list_field(row.get("linh_vuc_nghien_cuu"))
            for lv_name in linh_vucs:
                lv_name = lv_name.upper()
                conn.write("""
                    MATCH (gv:GiangVien) WHERE gv.id = $gv_id
                    MERGE (lv:LinhVucNghienCuu {ten_linh_vuc: $lv_name})
                    ON CREATE SET lv.id = 'lv_' + toString(id(lv))
                    MERGE (gv)-[:NGHIEN_CUU]->(lv)
                """, {"gv_id": gv_id, "lv_name": lv_name})

        except Exception as e:
            errors.append(f"Dòng {idx} ({ho_va_ten}): {str(e)}")

    return {"created": created, "updated": updated, "errors": errors}


def import_cong_trinh(df: pd.DataFrame, conn) -> dict:
    """
    Cột bắt buộc: ten_cong_trinh
    Cột tuỳ chọn: nam_xuat_ban, tom_tat, trang_thai, link,
                  tac_gia_giang_vien (email hoặc tên, phân cách '|'),
                  tac_gia_ngoai (tên, phân cách '|')
    """
    created, updated, errors = 0, 0, []
    rows = df.to_dict(orient="records")

    for idx, row in enumerate(rows, start=2):
        ten = safe_str(row.get("ten_cong_trinh"))
        if not ten:
            errors.append(f"Dòng {idx}: thiếu tên công trình – bỏ qua.")
            continue

        ten = " ".join(ten.split()).upper()
        ten_vi = " ".join(safe_str(row.get("ten_cong_trinh_vi")).split()).upper()
        nam_xuat_ban_str = safe_str(row.get("nam_xuat_ban"))
        nam_xuat_ban = int(nam_xuat_ban_str) if nam_xuat_ban_str.isdigit() else None

        slug = generate_slug(ten)
        slug_vi = generate_slug(ten_vi) if ten_vi else None

        props = {
            "ten_cong_trinh": ten,
            "ten_cong_trinh_vi": ten_vi if ten_vi else None,
            "slug":           slug,
            "slug_vi":        slug_vi,
            "nam_xuat_ban":   nam_xuat_ban,
            "noi_xuat_ban":   safe_str(row.get("noi_xuat_ban")).upper() if safe_str(row.get("noi_xuat_ban")) else None,
            "tom_tat":        safe_str(row.get("tom_tat")),
            "trang_thai":     safe_str(row.get("trang_thai")) or "Hoàn thành",
            "link":           safe_str(row.get("link")),
        }

        try:
            # Kiểm tra trùng theo slug tiếng Anh hoặc slug tiếng Việt
            existing = conn.query_single("""
                MATCH (ct:CongTrinhNghienCuu)
                WHERE (ct.slug = $slug AND coalesce(ct.is_deleted, false) = false)
                   OR ($slug_vi IS NOT NULL AND ct.slug_vi = $slug_vi AND coalesce(ct.is_deleted, false) = false)
                RETURN ct.id AS ct_id
            """, {"slug": slug, "slug_vi": slug_vi})

            if existing:
                errors.append(f"Dòng {idx}: Công trình '{ten}' đã tồn tại trong hệ thống (trùng lặp).")
                continue

            result = conn.write("""
                CREATE (ct:CongTrinhNghienCuu)
                SET ct.id = 'ct_' + toString(id(ct)),
                    ct.created_at = timestamp(),
                    ct.is_deleted = false,
                    ct += $props
                RETURN ct.id AS ct_id
            """, {"props": props})
            ct_id = result[0]["ct_id"]
            created += 1

            # Tác giả là Giảng viên
            tac_gia_gv = parse_list_field(row.get("tac_gia_giang_vien"))
            for idx_tg, tg in enumerate(tac_gia_gv):
                rel_type = "TAC_GIA_CHINH" if idx_tg == 0 else "CONG_SU"
                # Tìm theo email trước, nếu không có thì theo tên
                if "@" in tg:
                    conn.write(f"""
                        MATCH (gv:GiangVien {{email: $key}}),
                              (ct:CongTrinhNghienCuu {{id: $ct_id}})
                        WHERE coalesce(gv.is_deleted, false) = false
                        MERGE (gv)-[:{rel_type}]->(ct)
                    """, {"key": tg, "ct_id": ct_id})
                else:
                    conn.write(f"""
                        MATCH (gv:GiangVien), (ct:CongTrinhNghienCuu {{id: $ct_id}})
                        WHERE toUpper(gv.ho_va_ten) = toUpper($key)
                          AND coalesce(gv.is_deleted, false) = false
                        MERGE (gv)-[:{rel_type}]->(ct)
                    """, {"key": tg, "ct_id": ct_id})

            # Tác giả ngoài
            tac_gia_ngoai = parse_list_field(row.get("tac_gia_ngoai"))
            for ten_tgn in tac_gia_ngoai:
                ten_tgn = ten_tgn.upper()
                conn.write("""
                    MERGE (tgn:TacGiaNgoai {ho_va_ten: $ten})
                    ON CREATE SET tgn.id = 'tgn_' + toString(id(tgn)),
                                  tgn.hoc_vi = "",
                                  tgn.chuc_danh = "",
                                  tgn.don_vi_cong_tac = "",
                                  tgn.email = "",
                                  tgn.trang_thai = "Đã duyệt",
                                  tgn.is_deleted = false
                    WITH tgn
                    MATCH (ct:CongTrinhNghienCuu {id: $ct_id})
                    MERGE (tgn)-[:DONG_TAC_GIA]->(ct)
                """, {"ten": ten_tgn, "ct_id": ct_id})

            # Dọn dẹp quan hệ trùng lặp: nếu vừa là tác giả chính vừa là cộng sự/đồng tác giả
            conn.write("""
                MATCH (gv:GiangVien)-[:TAC_GIA_CHINH]->(ct:CongTrinhNghienCuu {id: $ct_id})
                MATCH (gv)-[r:CONG_SU|LA_TAC_GIA_CUA]->(ct)
                DELETE r
            """, {"ct_id": ct_id})

        except Exception as e:
            errors.append(f"Dòng {idx} ({ten}): {str(e)}")

    return {"created": created, "updated": 0, "errors": errors}


def import_de_tai(df: pd.DataFrame, conn) -> dict:
    """
    Cột bắt buộc: ten_de_tai
    Cột tuỳ chọn: cap_de_tai, nam_bat_dau, nam_ket_thuc, tom_tat,
                  trang_thai, link,
                  chu_nhiem (email/tên GV, phân cách '|'),
                  thanh_vien (email/tên GV, phân cách '|'),
                  tac_gia_ngoai (tên, phân cách '|')
    """
    created, updated, errors = 0, 0, []
    rows = df.to_dict(orient="records")

    for idx, row in enumerate(rows, start=2):
        ten = safe_str(row.get("ten_de_tai"))
        if not ten:
            errors.append(f"Dòng {idx}: thiếu tên đề tài – bỏ qua.")
            continue

        ten = " ".join(ten.split()).upper()
        nam_str = safe_str(row.get("nam")) or safe_str(row.get("nam_bat_dau")) or safe_str(row.get("nam_ket_thuc"))
        nam = int(nam_str) if nam_str.isdigit() else None

        props = {
            "ten_de_tai":   ten,
            "slug":         generate_slug(ten),
            "cap_de_tai":   safe_str(row.get("cap_de_tai")).upper() if safe_str(row.get("cap_de_tai")) else None,
            "nam":          nam,
            "tom_tat":      safe_str(row.get("tom_tat")),
            "trang_thai":   safe_str(row.get("trang_thai")) or "Đang thực hiện",
            "link":         safe_str(row.get("link")),
        }

        try:
            # Kiểm tra trùng đề tài
            slug = props["slug"]
            existing = conn.query_single("""
                MATCH (dt:DeTaiNghienCuu)
                WHERE (dt.slug = $slug OR toUpper(dt.ten_de_tai) = toUpper($ten_de_tai))
                  AND coalesce(dt.is_deleted, false) = false
                RETURN dt.id AS dt_id
            """, {"slug": slug, "ten_de_tai": ten})

            if existing:
                errors.append(f"Dòng {idx}: Đề tài '{ten}' đã tồn tại trong hệ thống (trùng lặp).")
                continue

            result = conn.write("""
                CREATE (dt:DeTaiNghienCuu)
                SET dt.id = 'dt_' + toString(id(dt)),
                    dt.slug = $slug,
                    dt.created_at = timestamp(),
                    dt.is_deleted = false,
                    dt += $props
                RETURN dt.id AS dt_id
            """, {"slug": slug, "props": props})

            dt_id = result[0]["dt_id"]
            created += 1

            def _link_gv(col: str, rel: str):
                for tg in parse_list_field(row.get(col)):
                    if "@" in tg:
                        conn.write(f"""
                            MATCH (gv:GiangVien {{email: $key}}),
                                  (dt:DeTaiNghienCuu {{id: $dt_id}})
                            WHERE coalesce(gv.is_deleted, false) = false
                            MERGE (gv)-[:{rel}]->(dt)
                        """, {"key": tg, "dt_id": dt_id})
                    else:
                        conn.write(f"""
                            MATCH (gv:GiangVien), (dt:DeTaiNghienCuu {{id: $dt_id}})
                            WHERE toUpper(gv.ho_va_ten) = toUpper($key)
                              AND coalesce(gv.is_deleted, false) = false
                            MERGE (gv)-[:{rel}]->(dt)
                        """, {"key": tg, "dt_id": dt_id})

            _link_gv("chu_nhiem", "CHU_NHIEM")
            _link_gv("thanh_vien", "THAM_GIA")

            for ten_tgn in parse_list_field(row.get("tac_gia_ngoai")):
                ten_tgn = ten_tgn.upper()
                conn.write("""
                    MERGE (tgn:TacGiaNgoai {ho_va_ten: $ten})
                    ON CREATE SET tgn.id = 'tgn_' + toString(id(tgn)),
                                  tgn.hoc_vi = "",
                                  tgn.chuc_danh = "",
                                  tgn.don_vi_cong_tac = "",
                                  tgn.email = "",
                                  tgn.trang_thai = "Đã duyệt",
                                  tgn.is_deleted = false
                    WITH tgn
                    MATCH (dt:DeTaiNghienCuu {id: $dt_id})
                    MERGE (tgn)-[:DONG_TAC_GIA]->(dt)
                """, {"ten": ten_tgn, "dt_id": dt_id})

            # Dọn dẹp quan hệ trùng lặp: nếu vừa là chủ nhiệm vừa là thành viên tham gia
            conn.write("""
                MATCH (gv:GiangVien)-[:CHU_NHIEM]->(dt:DeTaiNghienCuu {id: $dt_id})
                MATCH (gv)-[r:THAM_GIA]->(dt)
                DELETE r
            """, {"dt_id": dt_id})

        except Exception as e:
            errors.append(f"Dòng {idx} ({ten}): {str(e)}")

    return {"created": created, "updated": 0, "errors": errors}


def import_bo_mon(df: pd.DataFrame, conn) -> dict:
    """
    Cột bắt buộc: ten_bo_mon
    Cột tuỳ chọn: mo_ta, truong_bo_mon (email/tên GV)
    """
    created, updated, errors = 0, 0, []
    rows = df.to_dict(orient="records")

    for idx, row in enumerate(rows, start=2):
        ten = safe_str(row.get("ten_bo_mon"))
        if not ten:
            errors.append(f"Dòng {idx}: thiếu tên bộ môn – bỏ qua.")
            continue

        ten = ten.upper()
        props = {
            "ten_bo_mon": ten,
            "mo_ta":      safe_str(row.get("mo_ta")),
        }

        try:
            # Kiểm tra trùng bộ môn
            existing = conn.query_single("""
                MATCH (bm:BoMon)
                WHERE toUpper(bm.ten_bo_mon) = toUpper($ten_bo_mon)
                  AND coalesce(bm.is_deleted, false) = false
                RETURN bm.id AS bm_id
            """, {"ten_bo_mon": ten})

            if existing:
                errors.append(f"Dòng {idx}: Bộ môn '{ten}' đã tồn tại trong hệ thống (trùng lặp).")
                continue

            result = conn.write("""
                CREATE (bm:BoMon)
                SET bm.id = 'bm_' + toString(id(bm)),
                    bm.created_at = timestamp(),
                    bm.is_deleted = false,
                    bm += $props
                RETURN bm.id AS bm_id
            """, {"props": props})

            bm_id = result[0]["bm_id"]
            created += 1

            truong = safe_str(row.get("truong_bo_mon"))
            if truong:
                if "@" in truong:
                    conn.write(f"""
                        MATCH (gv:GiangVien {{email: $key}}),
                              (bm:BoMon {{id: $bm_id}})
                        WHERE coalesce(gv.is_deleted, false) = false
                        MERGE (gv)-[:TRUONG_BO_MON]->(bm)
                    """, {"key": truong, "bm_id": bm_id})
                else:
                    conn.write(f"""
                        MATCH (gv:GiangVien), (bm:BoMon {{id: $bm_id}})
                        WHERE toUpper(gv.ho_va_ten) = toUpper($key)
                          AND coalesce(gv.is_deleted, false) = false
                        MERGE (gv)-[:TRUONG_BO_MON]->(bm)
                    """, {"key": truong, "bm_id": bm_id})

        except Exception as e:
            errors.append(f"Dòng {idx} ({ten}): {str(e)}")

    return {"created": created, "updated": 0, "errors": errors}


# ─────────────────────────────────────────────
#  Dispatch map
# ─────────────────────────────────────────────
IMPORT_HANDLERS = {
    "giang-vien":  import_giang_vien,
    "cong-trinh":  import_cong_trinh,
    "de-tai":      import_de_tai,
    "bo-mon":      import_bo_mon,
}


# ─────────────────────────────────────────────
#  Endpoints
# ─────────────────────────────────────────────

@admin_import_bp.route("/import/upload", methods=["POST"])
def import_upload():
    """
    POST /api/admin/import/upload
    Form-data:
      - file      : file Excel hoặc CSV
      - data_type : 'giang-vien' | 'cong-trinh' | 'de-tai' | 'bo-mon'
    """
    if "file" not in request.files:
        return jsonify({"status": "error", "message": "Không tìm thấy file upload."}), 400

    file = request.files["file"]
    data_type = request.form.get("data_type", "").strip()

    if not file.filename:
        return jsonify({"status": "error", "message": "Tên file rỗng."}), 400

    ext = file.filename.rsplit(".", 1)[-1].lower()
    if ext not in ("xlsx", "xls", "csv"):
        return jsonify({"status": "error", "message": "Chỉ hỗ trợ file .xlsx, .xls hoặc .csv"}), 400

    if data_type not in IMPORT_HANDLERS:
        return jsonify({
            "status": "error",
            "message": f"Loại dữ liệu '{data_type}' không hợp lệ. "
                       f"Chọn một trong: {', '.join(IMPORT_HANDLERS.keys())}"
        }), 400

    try:
        df = read_file_to_df(file)
        if df.empty:
            return jsonify({"status": "error", "message": "File không có dữ liệu."}), 400

        total_rows = len(df)
        conn = get_neo4j_connection()
        result = IMPORT_HANDLERS[data_type](df, conn)

        # Nếu không có dòng nào mới được tạo/cập nhật (tất cả bị trùng hoặc lỗi), status nên là error
        if result["created"] == 0 and result.get("updated", 0) == 0:
            return jsonify({
                "status": "error",
                "message": "Dữ liệu đã tồn tại trong hệ thống (trùng lặp).",
                "total_rows": total_rows,
                "created":    0,
                "updated":    0,
                "error_count": len(result["errors"]),
                "errors":     result["errors"][:50],
            })

        return jsonify({
            "status": "ok",
            "message": "Import hoàn tất.",
            "total_rows": total_rows,
            "created":    result["created"],
            "updated":    result["updated"],
            "error_count": len(result["errors"]),
            "errors":     result["errors"][:50],   # tối đa 50 lỗi trả về
        })

    except Exception as e:
        return jsonify({"status": "error", "message": f"Lỗi xử lý file: {str(e)}"}), 500


@admin_import_bp.route("/import/template/<data_type>", methods=["GET"])
def download_template(data_type: str):
    """
    GET /api/admin/import/template/<data_type>
    Trả về file Excel mẫu để người dùng điền dữ liệu.
    """
    templates: dict[str, list[str]] = {
        "giang-vien": [
            "ho_va_ten", "ma_gv", "hoc_vi", "chuc_danh", "chuc_vu",
            "email", "dien_thoai", "chuyen_nganh", "trang_thai_cong_tac",
            "bo_mon", "linh_vuc_nghien_cuu"
        ],
        "cong-trinh": [
            "ten_cong_trinh", "ten_cong_trinh_vi", "nam_xuat_ban", "noi_xuat_ban",
            "tom_tat", "trang_thai", "link",
            "tac_gia_giang_vien", "tac_gia_ngoai"
        ],
        "de-tai": [
            "ten_de_tai", "cap_de_tai", "nam",
            "tom_tat", "trang_thai", "link",
            "chu_nhiem", "thanh_vien", "tac_gia_ngoai"
        ],
        "bo-mon": [
            "ten_bo_mon", "mo_ta", "truong_bo_mon"
        ],
    }

    if data_type not in templates:
        return jsonify({"status": "error", "message": "Loại không hợp lệ."}), 400

    columns = templates[data_type]
    df_template = pd.DataFrame(columns=columns)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:  # type: ignore[arg-type]
        df_template.to_excel(writer, index=False, sheet_name="Data")
        # Định dạng header
        ws = writer.sheets["Data"]
        from openpyxl.styles import Font, PatternFill, Alignment  # type: ignore[import-untyped]
        header_font  = Font(bold=True, color="FFFFFF")
        header_fill  = PatternFill("solid", fgColor="3B82F6")
        header_align = Alignment(horizontal="center", vertical="center")
        for cell in ws[1]:
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align
            ws.column_dimensions[cell.column_letter].width = max(len(str(cell.value)) + 6, 20)

    buf.seek(0)

    from flask import send_file
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"template_{data_type}.xlsx"
    )
